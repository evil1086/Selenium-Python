import openpyxl

book = openpyxl.load_workbook("C:\\Users\\vin\\Desktop\\testdata.xlsx")
sheet = book.active
dict = {}
cell = sheet.cell(row=1, column=2 )
print(cell.value)
# adding entry into the specified cell
sheet.cell(row=2, column=2).value = "Vinod"
print(sheet.cell(row=2, column=2).value)
# It will print the maximum count of rows and columns
print(sheet.max_column)
print(sheet.max_row)

# by using the Axis value
sheet["A5"].value = "Tester"
print(sheet["A5"].value)

for i in range(1, sheet.max_row + 1):
    # we are fetching values base on conditional
    if sheet.cell(row=i, column=1).value == "TestCase1" :
        for j in range(2, sheet.max_column + 1):
            dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(dict)