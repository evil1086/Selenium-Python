import openpyxl


class HomePageData:

    test_home_page_data = [{"FirstName":"Vinod", "E-mail":"vinod.pawar@maildrop.cc", "LastName":"pawar",  "Gender":"Male", "DOB":"09091995"},
                            {"FirstName":"Pallavi", "E-mail":"pallavi.pawar@maildrop.cc", "LastName":"Pawar", "Gender":"Female", "DOB":"26111995"}]

    @staticmethod
    def get_testdata(test_case_name):
        dict = {}
        book = openpyxl.load_workbook("C:\\Users\\vin\\Desktop\\testdata.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            # we are fetching values base on conditional
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return[dict]

